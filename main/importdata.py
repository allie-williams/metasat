# standard lib packages
import sys
import os
import requests
import json
import urllib.parse
import time
import csv
from openpyxl import Workbook

os.environ.setdefault('DJANGO_SETTINGS_MODULE','main.settings')

import django
django.setup()

from crosswalk.models import ExternalSchema, ExternalRelationshipSchema, Relationship, ExternalElement
from metasat.models import Element, Segment, ElementFamily

## set up databases

# def add_Segment(data1):
#     d, created = Segment.objects.get_or_create(segment=data1)
#     return d

# add_Segment("Space Segment")
# add_Segment("Ground Segment")
# add_Segment("Launch Segment")
# add_Segment("User Segment")


# def add_Family(data1):
#     d, created = ElementFamily.objects.get_or_create(family=data1)
#     return d

# add_Family("Attitude Control")
# add_Family("Communications")
# add_Family("Computer Hardware")
# add_Family("Data")
# add_Family("Electrical")
# add_Family("General")
# add_Family("Instrumentation")
# add_Family("Lens")
# add_Family("Mission")
# add_Family("Observation")
# add_Family("Optics")
# add_Family("Orbital Mechanics")
# add_Family("Product")
# add_Family("Propulsion")
# add_Family("Signal Processing")
# add_Family("Software")
# add_Family("Thermal Control")



# def add_SchemaType(data1):
#     d, created = SchemaType.objects.get_or_create(schematype=data1)
#     return d

## add_SchemaType("Boolean")
# add_SchemaType("text")
# add_SchemaType("integer")


def add_ExternalSchema(data1):
    d, created = ExternalSchema.objects.get_or_create(name=data1,url=data2)
    return d

# add_ExternalSchema("wikidata","http://wikidata.org")


def add_ExternalRelationshipSchema(data1, data2):
    d, created = ExternalRelationshipSchema.objects.get_or_create(name=data1,url=data2)
    return d

#add_ExternalRelationshipSchema("skos","http://skos.org")

def add_Relationship(data1, data2):
    d, created = Relationship.objects.get_or_create(predicate=data1,source_id=data2)
    return d

#add_Relationship("broader",2)



## add content to databases


def add_ExternalElement(data1, data2, data3, data4, data5, data6):
    d, created = ExternalElement.objects.get_or_create(identifier=data1, label=data2, url=data3, source_id=data4, relationship_id=data5, metasat_element_id=data6)

    return d


# def add_Element(data1, data2, data3, data4, data5, data6, segments):
#     #d, created = Element.objects.get_or_create(identifier=data1, desc=data2, synonym=data3, example=data4, schematype_id=data5)


#     e1 = Element(identifier=data1, desc=data2, synonym=data3, example=data4, schematype_id=data5, source=data6)
#     e1.save()

#     for x in segments:
#         e1.segment.add(x)
#     e1.save()


s1 = Segment.objects.get(id=1) # Space Segment
s2 = Segment.objects.get(id=2) # Ground Segment
s3 = Segment.objects.get(id=3) # Launch Segment
s4 = Segment.objects.get(id=4) # User Segment

f1 = ElementFamily.objects.get(id=1) # Attitude Control
f2 = ElementFamily.objects.get(id=2) # Communications
f3 = ElementFamily.objects.get(id=3) # Computer Hardware
f4 = ElementFamily.objects.get(id=4) # Data
f5 = ElementFamily.objects.get(id=5) # Electrical
f6 = ElementFamily.objects.get(id=6) # General
f7 = ElementFamily.objects.get(id=7) # Instrumentation
f8 = ElementFamily.objects.get(id=8) # Lens
f9 = ElementFamily.objects.get(id=9) # Mission
f10 = ElementFamily.objects.get(id=10) # Observation
f11 = ElementFamily.objects.get(id=11) # Optics
f12 = ElementFamily.objects.get(id=12) # Orbital Mechanics
f13 = ElementFamily.objects.get(id=13) # Product
f14 = ElementFamily.objects.get(id=14) # Propulsion
f15 = ElementFamily.objects.get(id=15) # Signal Processing
f16 = ElementFamily.objects.get(id=16) # Software
f17 = ElementFamily.objects.get(id=17) # Thermal Control



def add_Element(data1, data2, data3, data4, data5, data6, segments, families):
    #d, created = Element.objects.get_or_create(identifier=data1, desc=data2, synonym=data3, example=data4, schematype_id=data5)


    e1 = Element(identifier=data1, term=data2, desc=data3, synonym=data4, example=data5, source=data6)
    e1.save()

    for x in segments:
        e1.segment.add(x)

    for x in families:
        e1.family.add(x)
    e1.save()


#Element.objects.all().delete()


def update_Element(data1, data2, data3, data4, data5, data6, segments, families):
    d = Element.objects.get(identifier=data1)
    d.term = data2 
    d.desc = data3
    d.synonym = data4
    d.example = data5
    d.source = data6

    d.segment.clear()
    for x in segments:
        d.segment.add(x)

    d.family.clear()
    for x in families:
        d.family.add(x)
    d.save()

  
#add_Element("owner","Person or organization who owns the ground station","", "", 2, [s1,s2,s3])


from openpyxl import load_workbook

wb = load_workbook(filename = 'metasat6.xlsx')


#wb_obj = openpyxl.load_workbook(path) 
  
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb.active 
  
# Cell objects also have row, column,  
# and coordinate attributes that provide 
# location information for the cell. 
  
# Note: The first row or  
# column integer is 1, not 0. 
  
# Cell object is created by using  
# sheet object's cell() method. 
cell_obj = sheet_obj.cell(row = 2, column = 5) 
  
# Print value of cell object  
# using the value attribute 
#print(cell_obj.value) 

# sheet_ranges = wb['Term','Identifier']
# print(sheet_ranges['D18'].value)

num = sheet_obj.max_row

for x in range(1,num+1):

    term = sheet_obj.cell(row = x, column = 1) 
    identifier = sheet_obj.cell(row = x, column = 7) 
    families = sheet_obj.cell(row = x, column = 3) 
    segments = sheet_obj.cell(row = x, column = 4) 
    desc = sheet_obj.cell(row = x, column = 5) 
    example = sheet_obj.cell(row = x, column = 6)
    synonyms = sheet_obj.cell(row = x, column = 8)
    source = sheet_obj.cell(row = x, column = 10)


    familys = (families.value).split(", ")

    fam = []
    for x in familys:

        if x == "Attitude Control":
            fam.append(f1)
        if x == "Communications":
            fam.append(f2)
        if x == "Computer Hardware":
            fam.append(f3)
        if x == "Data":
            fam.append(f4)
        if x == "Electrical":
            fam.append(f5)
        if x == "General":
            fam.append(f6)
        if x == "Instrumentation":
            fam.append(f7)
        if x == "Lens":
            fam.append(f8)
        if x == "Mission":
            fam.append(f9)
        if x == "Observation":
            fam.append(f10)
        if x == "Optics":
            fam.append(f11)
        if x == "Orbital Mechanics":
            fam.append(f12)
        if x == "Product":
            fam.append(f13)
        if x == "Propulsion":
            fam.append(f14)
        if x == "Signal Processing":
            fam.append(f15)
        if x == "Software":
            fam.append(f16)
        if x == "Thermal Control":
            fam.append(f17)

    segs1 = (segments.value).split(", ")
    seg = []
    for x in segs1:
        if x == "Space Segment":
            seg.append(s1)
        if x == "Ground Segment":
            seg.append(s2)
        if x == "Launch Segment":
            seg.append(s3)
        if x == "User Segment":
            seg.append(s4)



    try:

        #e1 = Element(identifier=data1, term=data2, desc=data3, synonym=data4, example=data5, source=data6)
        add_Element(identifier.value,term.value,desc.value,synonyms.value,example.value,source.value, seg, fam)
    except django.db.utils.IntegrityError:
        print("dupe concept")
        print(identifier.value)
        update_Element(identifier.value,term.value,desc.value,synonyms.value,example.value,source.value, seg, fam)
        #pass

# with open('metasat2.csv') as csv_file:
#     csv_reader = csv.reader(csv_file, delimiter=',')
#     for row in csv_reader:
#         print(row)


#         add_Element(row[1],row[0],row[2],row[3],row[4],row[5])



# with open('crosswalk.csv') as csv_file:
#     csv_reader = csv.reader(csv_file, delimiter=',')
#     for row in csv_reader:
#         print(row)

        #add_ExternalElement("P127","Owned by","https://www.wikidata.org/wiki/Property:P127",1,1,15)
        #add_ExternalElement(name, barcode, notes)


#     print ("finished!")








# try:
#         familys = (families.value).split(", ")

#         if "Attitude Control" in familys:
#             familys[familys.index("Attitude Control")] = f1
#         if "Communications" in familys:
#             familys[familys.index("Communications")] = f2
#         if "Computer Hardware" in familys:
#             familys[familys.index("Computer Hardware")] = f3
#         if "Data" in familys:
#             familys[familys.index("Data")] = f4
#         if "Electrical" in familys:
#             familys[familys.index("Electrical")] = f5
#         if "General" in familys:
#             familys[familys.index("General")] = f6
#         if "Instrumentation" in familys:
#             familys[familys.index("Instrumentation")] = f7
#         if "Lens" in familys:
#             familys[familys.index("Lens")] = f8
#         if "Mission" in familys:
#             familys[familys.index("Mission")] = f9
#         if "Observation" in familys:
#             familys[familys.index("Observation")] = f10
#         if "Optics" in familys:
#             familys[familys.index("Optics")] = f11
#         if "Orbital Maneuver" in familys:
#             familys[familys.index("Orbital Maneuver")] = f12
#         if "Orbital Mechanics" in familys:
#             familys[familys.index("Orbital Mechanics")] = f13
#         if "Product" in familys:
#             familys[familys.index("Product")] = f14
#         if "Propulsion" in familys:
#             familys[familys.index("Propulsion")] = f15
#         if "Signal Processing" in familys:
#             familys[familys.index("Signal Processing")] = f16
#         if "Software" in familys:
#             familys[familys.index("Software")] = f17
#         if "Thermal Control" in familys:
#             familys[familys.index("Thermal Control")] = f18

#         print(familys)


#         segment = (segments.value).split(", ")
        
#         if "Space Segment" in segment:
#             segment[segment.index("Space Segment")] = s1
#         if "Ground Segment" in segment:
#             segment[segment.index("Ground Segment")] = s2
#         if "Launch Segment" in segment:
#             segment[segment.index("Launch Segment")] = s3
#         if "User Segment" in segment:
#             segment[segment.index("User Segment")] = s4